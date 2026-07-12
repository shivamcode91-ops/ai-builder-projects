"""Deviation tests: reviewers may feed datarooms that do not match the three
provided companies byte-for-byte. These pin the hardened behaviour — parse by
name not position, tolerate format drift, and degrade instead of crashing."""
from datetime import datetime

import openpyxl
import pytest

from pipeline.ingest.excel import ingest_excel, parse_bank_statement
from pipeline.ingest import docs as docmod


def _write_statement(path, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ACME BANK LTD"])          # bank letterhead above the header
    ws.append([])
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_bank_statement_deviated_layout(tmp_path):
    """Renamed headers, reordered columns, dd/mm/yyyy dates, 'ECS RTN' bounce
    wording — everything the original parser would have crashed on."""
    p = tmp_path / "Bank_Statement_FY27.xlsx"
    _write_statement(
        p,
        ["Transaction Date", "Narration", "Ref No", "Deposit Amt", "Withdrawal Amt", "Running Balance"],
        [
            ["01/04/2026", "NEFT INWARD - CUSTOMER A", "r1", 500000, None, 500000],
            ["15/04/2026", "VENDOR PAYMENT", "r2", None, 200000, 300000],
            ["03/05/2026", "ECS RTN - INSUFFICIENT FUNDS", "r3", None, 45000, 255000],
            ["03/05/2026", "ECS RTN CHARGES", "r4", None, 590, 254410],
            ["20/05/2026", "NEFT INWARD - CUSTOMER B", "r5", 100000, None, 354410],
            ["TOTAL", "", "", 600000, 245590, None],   # footer row must be skipped
        ],
    )
    out = parse_bank_statement(p)
    assert out["n_transactions"] == 5
    assert out["bounce_count_12m"] == 1                       # RTN counted, fee row excluded
    assert out["monthly"]["Apr-26"]["inflows"] == 500000
    assert out["monthly"]["Apr-26"]["outflows"] == 200000
    assert out["monthly"]["May-26"]["inflows"] == 100000
    assert out["closing_balance_inr"] == 354410


def test_bank_statement_native_datetime_cells(tmp_path):
    """Excel often stores dates as real datetimes, not strings."""
    p = tmp_path / "Bank_Statement.xlsx"
    _write_statement(
        p,
        ["Txn Date", "Ref", "Description", "Cheque", "Debit", "Credit", "Balance"],
        [[datetime(2026, 4, 1), "r", "NEFT INWARD", None, None, 250000, 250000]],
    )
    out = parse_bank_statement(p)
    assert out["n_transactions"] == 1
    assert out["monthly"]["Apr-26"]["inflows"] == 250000


def test_bank_statement_no_header_raises_clearly(tmp_path):
    p = tmp_path / "Bank_Statement.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["just", "some", "cells"])
    wb.save(p)
    with pytest.raises(ValueError, match="no header row"):
        parse_bank_statement(p)


def test_ingest_isolates_corrupt_file(tmp_path):
    """A corrupt xlsx becomes a parse_warning; the other parsers still run."""
    (tmp_path / "Bank_Statement_FY26.xlsx").write_bytes(b"this is not a workbook")
    good = tmp_path / "Cash_Position.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["Month", "Closing"])
    wb.active.append(["Apr-26", 1000000])
    wb.save(good)

    data = ingest_excel(tmp_path)
    assert "bank_statement" not in data
    assert "cash_position" in data                            # unaffected parser ran
    assert any("bank_statement" in w for w in data["parse_warnings"])


class _FakeLLM:
    """Returns canned responses in order — first junk, then valid JSON."""
    def __init__(self, answers):
        self.answers = list(answers)
        self.calls = 0

    def complete(self, name, model, messages, max_tokens=None, temperature=0.0):
        self.calls += 1
        return self.answers.pop(0)


def test_extraction_retries_on_invalid_json(tmp_path):
    png = tmp_path / "Sanction_Letter_scan.png"
    png.write_bytes(b"\x89PNG fake image bytes")
    llm = _FakeLLM(["Sure! Here is what I found in the letter.",
                    '{"lender": "ICICI Bank", "facility_type": "Cash Credit"}'])
    result = docmod.extract_document(llm, "sanction_letter", png, model="test-model")
    assert llm.calls == 2
    assert result["lender"] == "ICICI Bank"
    assert result["_source"] == png.name


def test_validator_skips_self_liquidating_products():
    """'no EMI obligation' must not trip the amortising-DSCR check."""
    from pipeline.structure import validate_structure
    snapshot = {"financials": {"ebitda_inr": {"FY26": 50_000_000}},
                "ratios_computed": {"annual_debt_service_inr": 88_000_000},
                "receivables": {"eligible_0_60_inr": 87_000_000}}
    structure = {"eligibility": "conditional", "rationale": "x",
                 "recommended_products": [{
                     "product": "Invoice Discounting", "amount_inr": 52_000_000,
                     "tenor_months": 3, "pricing": {"all_in_rate_pct": 11.0},
                     "repayment": "self-liquidating — no EMI obligation added",
                     "advance_rate_pct": 60}]}
    warnings = validate_structure(structure, snapshot)
    assert not any("DSCR" in w for w in warnings)
