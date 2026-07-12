"""Deterministic Excel ingestion — zero LLM tokens.

Everything numeric in the Dataroom lives in .xlsx files, so we parse and
compute it in code rather than asking a model to read spreadsheets. This is
cheaper, faster, and exact. The large files (bank statement ~2,700 rows, MIS
23 tabs) are reduced to compact aggregates here and never enter a prompt.
"""
import re
from datetime import datetime
from pathlib import Path

import openpyxl


def _rows(ws):
    return [r for r in ws.iter_rows(values_only=True)]


def _num(v, default=0.0):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("%", "").strip()
        try:
            return float(s)
        except ValueError:
            return default
    return default


def _find_file(root: Path, patterns):
    for pat in patterns:
        hits = sorted(root.rglob(pat))
        if hits:
            return hits[0]
    return None


# ---------------------------------------------------------------- financial model

def parse_financial_model(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"source": path.name}

    rows = _rows(wb["Assumptions"])
    out["assumptions"] = {str(r[0]).strip(): r[1] for r in rows if r[0] is not None}

    rows = _rows(wb["P&L"])
    years = [str(y) for y in rows[0][1:] if y]
    pnl = {}
    for r in rows[1:]:
        if r[0] is None:
            continue
        pnl[str(r[0]).strip()] = {years[i]: _num(r[i + 1]) for i in range(len(years))}
    out["pnl"] = pnl
    out["pnl_years"] = years

    rows = _rows(wb["Balance Sheet"])
    out["balance_sheet"] = {str(r[0]).strip(): _num(r[1]) for r in rows[1:] if r[0] is not None}

    rows = _rows(wb["Monthly Cashflow"])
    cf = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        cf.append({
            "month": str(r[0]), "opening": _num(r[1]), "inflows": _num(r[2]),
            "outflows": _num(r[3]), "closing": _num(r[4]), "bounces": int(_num(r[5])),
        })
    out["monthly_cashflow"] = cf

    rows = _rows(wb["Ratios"])
    out["ratios"] = {str(r[0]).strip(): _num(r[1]) for r in rows[1:] if r[0] is not None}
    return out


# ---------------------------------------------------------------- debt schedule

def parse_debt_schedule(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _rows(wb.active)
    facilities, totals = [], {}
    for r in rows[1:]:
        if r[0] is None:
            continue
        if str(r[0]).strip().upper() == "TOTAL":
            totals = {"sanctioned_inr": _num(r[3]), "outstanding_inr": _num(r[4]),
                      "monthly_emi_inr": _num(r[6])}
            continue
        facilities.append({
            "lender": str(r[0]).strip(), "facility_type": str(r[1]).strip(),
            "sanctioned_inr": _num(r[3]), "outstanding_inr": _num(r[4]),
            "roi_pct": _num(r[5]), "monthly_emi_inr": _num(r[6]),
            "tenor_months": int(_num(r[7])),
        })
    if not totals:
        totals = {
            "sanctioned_inr": sum(f["sanctioned_inr"] for f in facilities),
            "outstanding_inr": sum(f["outstanding_inr"] for f in facilities),
            "monthly_emi_inr": sum(f["monthly_emi_inr"] for f in facilities),
        }
    wc_types = ("CASH CREDIT", "WCDL", "OD", "OVERDRAFT")
    wc = [f for f in facilities if f["facility_type"].upper() in wc_types]
    totals["wc_sanctioned_inr"] = sum(f["sanctioned_inr"] for f in wc)
    totals["wc_outstanding_inr"] = sum(f["outstanding_inr"] for f in wc)
    totals["wc_utilisation_pct"] = round(
        100 * totals["wc_outstanding_inr"] / totals["wc_sanctioned_inr"], 1
    ) if totals["wc_sanctioned_inr"] else None
    return {"source": path.name, "facilities": facilities, "totals": totals}


# ---------------------------------------------------------------- debtor ageing

def parse_debtor_ageing(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _rows(wb.active)
    buckets, total = {}, 0.0
    for r in rows[1:]:
        if r[0] is None:
            continue
        label = str(r[0]).strip()
        if label.upper() == "TOTAL":
            total = _num(r[1])
            continue
        buckets[label] = _num(r[1])
    if not total:
        total = sum(buckets.values())
    eligible = sum(v for k, v in buckets.items() if k.startswith(("0-30", "31-60")))
    overdue_90 = sum(v for k, v in buckets.items() if k.startswith("90"))
    return {
        "source": path.name, "buckets_inr": buckets, "total_inr": total,
        "eligible_0_60_inr": eligible,
        "pct_90_plus": round(100 * overdue_90 / total, 1) if total else 0.0,
    }


# ---------------------------------------------------------------- cash position

def parse_cash_position(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _rows(wb.active)
    months = [{"month": str(r[0]), "closing_inr": _num(r[1])} for r in rows[1:] if r[0]]
    return {"source": path.name, "months": months}


# ---------------------------------------------------------------- bank statement

_DATE_FORMATS = ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d.%m.%Y")
# a return event, minus the paired fee row ('... CHARGES'/'FEE') that would double-count it
_BOUNCE_RE = re.compile(r"\b(RETURN|RETURNED|RTN|BOUNCE|BOUNCED)\b")
_FEE_RE = re.compile(r"\b(CHARGE|CHARGES|FEE|FEES)\b")


def _parse_date(v):
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _cell(row, j):
    return row[j] if j is not None and j < len(row) else None


def parse_bank_statement(path: Path):
    """~2,700 txn rows -> monthly aggregates + bounce events + closing balance.

    Bounces appear as two rows each ('CHEQUE RETURN - <reason>' and a separate
    'CHEQUE RETURN CHARGES' fee row); we count only the return itself.

    The header row and columns are resolved by NAME (with positional fallback)
    and dates accept common Indian bank formats plus native Excel datetimes, so
    a reformatted statement degrades gracefully instead of crashing.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = _rows(wb.active)

    def _txt(v):
        return str(v).strip().lower() if v is not None else ""

    header_idx, header = None, None
    for i, r in enumerate(rows[:50]):
        cells = [_txt(c) for c in r]
        if any("date" in c for c in cells) and any(
                k in c for c in cells for k in ("debit", "credit", "withdraw", "deposit")):
            header_idx, header = i, cells
            break
    if header_idx is None:
        raise ValueError(f"{path.name}: no header row with a date + debit/credit "
                         "column found in the first 50 rows")

    def _col(*keys, default=None):
        for j, c in enumerate(header):
            if any(k in c for k in keys):
                return j
        return default

    c_date = _col("date", default=0)
    c_desc = _col("description", "narration", "particular", default=2)
    c_debit = _col("debit", "withdraw", default=4)
    c_credit = _col("credit", "deposit", default=5)
    c_bal = _col("balance", default=6)

    monthly = {}
    bounces = []
    last_balance, last_date = None, None
    n_txns = 0
    for r in rows[header_idx + 1:]:
        raw_date = _cell(r, c_date)
        if raw_date is None:
            continue
        date = _parse_date(raw_date)
        if date is None:  # footer/summary rows
            continue
        n_txns += 1
        key = date.strftime("%b-%y")
        m = monthly.setdefault(key, {"inflows": 0.0, "outflows": 0.0, "txns": 0})
        debit, credit = _num(_cell(r, c_debit)), _num(_cell(r, c_credit))
        m["outflows"] += debit
        m["inflows"] += credit
        m["txns"] += 1
        desc = str(_cell(r, c_desc) or "").upper()
        if _BOUNCE_RE.search(desc) and not _FEE_RE.search(desc):
            bounces.append({"date": str(raw_date), "description": str(_cell(r, c_desc)),
                            "amount_inr": debit or credit})
        last_balance, last_date = _num(_cell(r, c_bal)), str(raw_date)
    n_months = len(monthly) or 1
    total_in = sum(m["inflows"] for m in monthly.values())
    total_out = sum(m["outflows"] for m in monthly.values())
    return {
        "source": path.name,
        "n_transactions": n_txns,
        "monthly": {k: {kk: round(vv, 2) for kk, vv in v.items()} for k, v in monthly.items()},
        "avg_monthly_inflows_inr": round(total_in / n_months, 0),
        "avg_monthly_outflows_inr": round(total_out / n_months, 0),
        "closing_balance_inr": last_balance,
        "closing_date": last_date,
        "bounce_count_12m": len(bounces),
        "bounce_events": bounces,
    }


# ---------------------------------------------------------------- MIS (optional)

def parse_mis(path: Path):
    """Company MIS workbook (23 tabs, registers up to ~2,150 rows) -> compact
    aggregates. Only present for some companies."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"source": path.name, "tabs": wb.sheetnames}

    if "Bank Facility Utilisation" in wb.sheetnames:
        rows = _rows(wb["Bank Facility Utilisation"])
        facs = []
        for r in rows[1:]:
            if r[0] is None or str(r[0]).strip().upper() == "TOTAL":
                continue
            facs.append({
                "lender": r[0], "facility": r[1], "sanctioned_inr": _num(r[2]),
                "drawing_power_inr": _num(r[3]), "outstanding_inr": _num(r[4]),
                "utilisation_pct": _num(r[5]), "roi_pct": _num(r[6]),
            })
        out["facility_utilisation"] = facs

    if "Working Capital" in wb.sheetnames:
        rows = _rows(wb["Working Capital"])
        last = rows[-1]
        vals = [r for r in rows[1:] if r[0]]
        out["working_capital"] = {
            "latest_month": str(last[0]), "dso_days": _num(last[5]),
            "dio_days": _num(last[6]), "dpo_days": _num(last[7]),
            "wc_cycle_days": _num(last[8]),
            "avg_wc_cycle_days": round(sum(_num(r[8]) for r in vals) / len(vals), 0),
        }

    if "Sales by Customer" in wb.sheetnames:
        rows = _rows(wb["Sales by Customer"])
        custs = [r for r in rows[1:] if r[0] and str(r[0]).strip().upper() != "TOTAL"]
        custs.sort(key=lambda r: _num(r[2]), reverse=True)
        total_sales = sum(_num(r[2]) for r in custs)
        out["customers"] = {
            "count": len(custs),
            "total_sales_inr": total_sales,
            "top5": [{"name": r[0], "region": r[1], "sales_inr": _num(r[2]),
                      "pct": round(100 * _num(r[2]) / total_sales, 1),
                      "outstanding_inr": _num(r[4])} for r in custs[:5]],
            "top1_pct": round(100 * _num(custs[0][2]) / total_sales, 1) if custs else None,
            "top5_pct": round(100 * sum(_num(r[2]) for r in custs[:5]) / total_sales, 1) if custs else None,
        }

    if "Debtor Ageing" in wb.sheetnames:
        rows = _rows(wb["Debtor Ageing"])
        total_row = next((r for r in rows if r[0] and str(r[0]).strip().upper() == "TOTAL"), None)
        if total_row:
            total = _num(total_row[5])
            out["debtor_ageing"] = {
                "buckets_inr": {"0-30": _num(total_row[1]), "31-60": _num(total_row[2]),
                                "61-90": _num(total_row[3]), "90+": _num(total_row[4])},
                "total_inr": total,
                "pct_90_plus": round(100 * _num(total_row[4]) / total, 1) if total else 0,
            }

    if "Order Book" in wb.sheetnames:
        rows = _rows(wb["Order Book"])
        orders = [r for r in rows[1:] if r[0]]
        confirmed = [r for r in orders if str(r[5]).strip().lower() == "confirmed"]
        out["order_book"] = {
            "orders": len(orders),
            "total_value_inr": sum(_num(r[3]) for r in orders),
            "confirmed_value_inr": sum(_num(r[3]) for r in confirmed),
        }
    return out


# ---------------------------------------------------------------- entry point

def ingest_excel(company_dir: Path):
    """One malformed sheet must not kill the run: each parser is isolated, and
    failures are recorded as parse_warnings that flow into the snapshot (so the
    structurer and the UI see what is missing instead of a stack trace)."""
    company_dir = Path(company_dir)
    data = {}
    warnings = []
    sources = [
        ("financial_model", ["Financial_Model*.xlsx"], parse_financial_model),
        ("debt_schedule", ["Debt_Schedule*.xlsx"], parse_debt_schedule),
        ("debtor_ageing", ["Debtor_Ageing*.xlsx"], parse_debtor_ageing),
        ("cash_position", ["Cash_Position*.xlsx"], parse_cash_position),
        ("bank_statement", ["Bank_Statement*.xlsx"], parse_bank_statement),
        ("mis", ["*MIS*.xlsx"], parse_mis),
    ]
    for key, patterns, parser in sources:
        path = _find_file(company_dir, patterns)
        if not path:
            continue
        try:
            data[key] = parser(path)
        except Exception as e:
            warnings.append(f"{key}: failed to parse {path.name} — {type(e).__name__}: {e}")
    if warnings:
        data["parse_warnings"] = warnings
    return data
